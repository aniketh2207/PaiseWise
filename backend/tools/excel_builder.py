import io
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import PieChart, LineChart, Reference

try:
    from backend.database.db import SessionLocal
    from backend.database.models import BankTransaction, MonthlySummary
except ModuleNotFoundError:
    from database.db import SessionLocal
    from database.models import BankTransaction, MonthlySummary


HEADER_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def build_excel_report(month: int, year: int) -> bytes:
    """
    Builds the full multi-sheet Excel report for a given month/year.
    Returns raw bytes — caller decides whether to save to disk or attach to email.
    """
    db = SessionLocal()
    try:
        summary = db.query(MonthlySummary).filter(
            MonthlySummary.month == month, MonthlySummary.year == year
        ).first()
        txns = db.query(BankTransaction).filter(
            BankTransaction.month == month, BankTransaction.year == year
        ).order_by(BankTransaction.date.desc()).all()

        if not summary:
            raise ValueError(f"No summary found for {month}/{year}. Run reconciliation first.")

        wb = Workbook()

        _build_summary_sheet(wb, summary, month, year)
        _build_category_sheet(wb, summary)
        _build_transactions_sheet(wb, txns)
        _build_daily_trend_sheet(wb, txns)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    finally:
        db.close()


def _build_summary_sheet(wb, summary, month, year):
    ws = wb.active
    ws.title = "Summary"

    month_name = datetime(year, month, 1).strftime("%B %Y")
    rows = [
        ("Period", month_name),
        ("Total Spent", f"₹{summary.total_debits:,.2f}"),
        ("Total Received", f"₹{summary.total_credits:,.2f}"),
        ("Net", f"₹{(summary.total_credits - summary.total_debits):,.2f}"),
        ("Top Category", summary.top_merchant or "N/A"),
        ("Transactions Tracked", summary.bank_txn_count or 0),
        ("Match Rate", f"{(summary.match_rate or 0) * 100:.0f}%"),
    ]

    ws["A1"] = "PaiseWise Monthly Report"
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells("A1:B1")

    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    if summary.llm_insights:
        ws.cell(row=len(rows) + 5, column=1, value="AI Insights").font = Font(bold=True, size=12)
        ws.cell(row=len(rows) + 6, column=1, value=summary.llm_insights)
        ws.merge_cells(start_row=len(rows) + 6, start_column=1, end_row=len(rows) + 6, end_column=4)
        ws.cell(row=len(rows) + 6, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40


def _build_category_sheet(wb, summary):
    ws = wb.create_sheet("Category Breakdown")
    by_category = json.loads(summary.by_category or "{}")

    ws.cell(row=1, column=1, value="Category").font = HEADER_FONT
    ws.cell(row=1, column=2, value="Amount").font = HEADER_FONT
    for c in (1, 2):
        ws.cell(row=1, column=c).fill = HEADER_FILL

    sorted_items = sorted(by_category.items(), key=lambda x: x[1], reverse=True)
    for i, (cat, amount) in enumerate(sorted_items, start=2):
        ws.cell(row=i, column=1, value=cat)
        ws.cell(row=i, column=2, value=round(amount, 2))

    if sorted_items:
        chart = PieChart()
        chart.title = "Spending by Category"
        data = Reference(ws, min_col=2, min_row=1, max_row=len(sorted_items) + 1)
        labels = Reference(ws, min_col=1, min_row=2, max_row=len(sorted_items) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        ws.add_chart(chart, "D2")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15


def _build_transactions_sheet(wb, txns):
    ws = wb.create_sheet("All Transactions")
    headers = ["Date", "Recipient/Merchant", "Reason", "Amount", "Type", "Category", "Subcategory", "Source"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    source_map = {
        "matched": "Slack ✓",
        "auto_categorized": "Auto-detected",
        "user_annotated": "Annotated",
        "needs_annotation": "Unmatched",
    }

    for i, t in enumerate(txns, start=2):
        ws.cell(row=i, column=1, value=t.date.isoformat())
        ws.cell(row=i, column=2, value=t.upi_name or "")
        ws.cell(row=i, column=3, value=t.reason or "")
        ws.cell(row=i, column=4, value=t.amount)
        ws.cell(row=i, column=5, value=t.type)
        ws.cell(row=i, column=6, value=t.category or "Uncategorized")
        ws.cell(row=i, column=7, value=t.subcategory or "")
        ws.cell(row=i, column=8, value=source_map.get(t.reconcile_status, t.reconcile_status or ""))

        row_fill = None
        if t.reconcile_status == "matched":
            row_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        elif t.reconcile_status == "auto_categorized":
            row_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        elif t.reconcile_status == "user_annotated":
            row_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
        elif t.reconcile_status == "needs_annotation":
            row_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

        if row_fill:
            for c in range(1, 9):
                ws.cell(row=i, column=c).fill = row_fill

    for col, width in zip("ABCDEFGH", [12, 22, 24, 12, 10, 16, 16, 14]):
        ws.column_dimensions[col].width = width


def _build_daily_trend_sheet(wb, txns):
    ws = wb.create_sheet("Daily Trend")
    ws.cell(row=1, column=1, value="Date").font = HEADER_FONT
    ws.cell(row=1, column=2, value="Amount Spent").font = HEADER_FONT

    daily = {}
    for t in txns:
        if t.type == "debit":
            daily[t.date] = daily.get(t.date, 0) + t.amount

    sorted_days = sorted(daily.items())
    for i, (d, amt) in enumerate(sorted_days, start=2):
        ws.cell(row=i, column=1, value=d.isoformat())
        ws.cell(row=i, column=2, value=round(amt, 2))

    if sorted_days:
        chart = LineChart()
        chart.title = "Daily Spending Trend"
        data = Reference(ws, min_col=2, min_row=1, max_row=len(sorted_days) + 1)
        labels = Reference(ws, min_col=1, min_row=2, max_row=len(sorted_days) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        ws.add_chart(chart, "D2")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16